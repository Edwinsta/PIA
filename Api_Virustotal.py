from virus_total_apis import PublicApi
from datetime import date
from openpyxl import Workbook
import time

# Primero introducimos la apikey y abrimos el archivo txt para despues leerlo
Api_key = "290aaeaf4e256b72789c44418b45e15a6e5506b4f9777ef1ef983ed398847b2f"
api = PublicApi(Api_key)
x = open("urls_sospechosas.txt", "r")
lines = x.readlines()
x.close()

#Despues se creo el Excel y seguido a eso le pusimos nombre a la hoja que usaremos y se elimina la hoja de sheet
libro = Workbook()
hoja = libro.active
hoja1 = libro.create_sheet("Labpro5")
libro.remove_sheet(libro.get_sheet_by_name("Sheet"))

# Lo siguiente es asignar variables a las celdas
linum1 = 2
v1 = hoja1['A1']
v1.value = "URL"

linum2 = 2
v2 = hoja1['B1']
v2.value = "Fecha de análisis"

linum3 = 2
v3 = hoja1['C1']
v3.value = "Total de análisis"

linum4 = 2
v4 = hoja1['D1']
v4.value = "Análisis positivos"

linum5 = 2
v5 = hoja1['E1']
v5.value = "Clasificación"

# Por consiguiente usamos un for para asi leer cada linea del archivo .txt y ejecutar el analisis
for url in lines:
    url = url.replace("\n", "")
    hoja1.cell(linum1, 1, url)
    time.sleep(15)
    response = api.get_url_report(url)
    linum1 += 1
    today = date.today()
    d3 = today.strftime("%d/%m/%Y")
    hoja1.cell(linum2, 2, d3)
    linum2 += 1
    if response["response_code"] == 200:
        hoja1.cell(linum3, 4, (response["results"]["positives"]))
        linum3 += 1
        if (response["results"]["positives"]) <= 3:
            hoja1.cell(linum4, 5, "Clasificación Baja")
            linum4 += 1
        elif (response["results"]["positives"]) <= 10:
            hoja1.cell(linum4, 5, "Clasificación Media")
            linum4 += 1
        else:
            hoja1.cell(linum4, 5, "Alta")
            linum4 += 1
        hoja1.cell(linum5, 3, (response["results"]["total"]))
        linum5 += 1
    else:
        print("El analidis del archivo no se ah podido efectuar.")


libro.save("reporte_analizador_urls.xlsx")
